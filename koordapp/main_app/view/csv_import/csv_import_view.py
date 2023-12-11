import string, random
from datetime import datetime
from django.views.generic import TemplateView
from django.views.generic.edit import CreateView
from django.contrib.auth import login as auth_login, authenticate, logout as auth_logout
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import LoginView, LogoutView
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.models import User, Group
from django.contrib import messages

from django.shortcuts import redirect
from django.shortcuts import render
from django.http import HttpResponse

from openpyxl import Workbook, load_workbook

from main_app.models import Nutzer, Personal, Raum, Gruppe, AG, Schueler 

def csv_import_view(request):
    if request.method == 'POST':
        excel_file=None
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
        response['Content-Disposition'] = 'attachment; filename={date}-otps.xlsx'.format(date=datetime.now().strftime('%Y-%m-%d'),)

        try:
            excel_file = request.FILES['excel_file']
        except:
            print("keine Datei vorhanden")
            return redirect('home')
        
        if(excel_file.name.endswith(".xlsx")):              # check if xlsx datei

            try:
                    wb = load_workbook(excel_file, read_only=True)  # save workbook as Excel Workbook (*.xlsx) and not as Strict Open XML Spreadsheet (*.xlsx)
                             # Parameter setzten
                    
                    fehler_tabelle=''
                    optionlist_reset = request.POST.getlist('checkbox_database')

                    optionlist = request.POST.getlist('checkbox_sheet')

                    if 'option_room' in optionlist:
                        if 'option_overwrite' in optionlist_reset:
                            Raum.objects.all().delete()

                        wss = wb['Raeume']
                        for row in wss.iter_rows(min_row=2, values_only=True):              # Erstellung Nutzer
                            raum_nr, geschoss, kapazitaet = row
                            if(Raum.objects.filter(raum_nr=raum_nr).exists()==False):
                                Raum.objects.create(raum_nr=raum_nr,geschoss=geschoss,kapazitaet=kapazitaet)

                    if 'option_user' in optionlist:
                        fehler_tabelle='Fehler in Tabelle Personal: '
                        wb_otp = Workbook()
                        activ_sheet = wb_otp.active
                        activ_sheet.title = "OTP"
                        activ_sheet.append(['Username', 'Einmal Passwort'])
                        wsp = wb['Personal']
                        for row in wsp.iter_rows(min_row=2, values_only=True):              # Erstellung Personal
                            
                            error = False
                            vorname, nachname, funktion, rechte = row
                            new_nutzer = Nutzer.objects.create(vorname=vorname,nachname=nachname)
                            try:
                                if rechte=='':
                                    rechte = 'Ohne Rolle'
                                rechte_gruppe = Group.objects.get(name=rechte)
                            except:
                                error=True
                                messages.error(request, fehler_tabelle+"rechte_gruppe "+rechte+" vom Personal "+vorname+" "+nachname+" in Zeile " + row.count +" existiert nicht. Options: (Admin, Gruppenleitung, Raumbetreuer, Ohne Rolle)")
                            randompw = ''.join(random.choice(string.ascii_letters+string.digits) for _ in range(6))     #erstellung random passwort
                            #print(randompw)
                            username = (vorname+nachname).lower()
                            zahl = 0
                            is_username_unique = False
                            while not is_username_unique:
                                un2 = username
                                if(zahl > 0):
                                    un2 = username + str(zahl)
                                if (User.objects.filter(username=un2).exists()==True):
                                    zahl += 1
                                else:
                                    username=un2
                                    is_username_unique=True
                            #print(username)
                            activ_sheet.append([username, randompw])
                            newuser = User.objects.create_user(username=username, password=randompw)
                            newuser.groups.add(rechte_gruppe)
                            if not error:
                                Personal.objects.create(rolle=funktion, nutzer=new_nutzer, user=newuser, rechte_gruppe=rechte_gruppe)
    
                        wb_otp.save(response)

                    if 'option_ag' in optionlist:
                        fehler_tabelle='Fehler in Tabelle AGs: '
                        if 'option_overwrite' in optionlist_reset:
                            AG.objects.all().delete()
                        wss = wb['AGs']
                        for row in wss.iter_rows(min_row=2, values_only=True):              # Erstellung Nutzer
                            error = False
                            name, beschreibung, max_anzahl, offene_AG, ag_leiter = row
                            zahl = 1
                            is_name_unique = False
                            while not is_name_unique:
                                un2 = name
                                if(zahl > 1):
                                    un2 = name + str(zahl)
                                if (AG.objects.filter(name=un2).exists()==True):
                                    zahl += 1
                                else:
                                    name=un2
                                    is_name_unique=True
                            try:
                                leiter = Personal.objects.get(user=(User.objects.get(username=ag_leiter)))
                            except:
                                error=True
                                messages.error(request, fehler_tabelle+"ag_leiter "+name+" in Zeile " + row.count +" existiert nicht.")
                            is_offene_AG = False
                            try:
                                max_anzahl = int(max_anzahl)
                            except:
                                error=True
                                messages.error(request, fehler_tabelle+"max_anzahl "+max_anzahl+" in Zeile " + row.count +" muss eine Zahl sein")
                            if(offene_AG.lower()=='true' or offene_AG.lower()=='ja'):
                                    is_offene_AG = True
                            if not error:        
                                AG.objects.create(name=name,beschreibung=beschreibung,max_anzahl=max_anzahl,offene_AG=is_offene_AG, leiter=leiter)
                    
                    if 'option_group' in optionlist:
                        fehler_tabelle='Fehler in Tabelle Gruppen: '
                        if 'option_overwrite' in optionlist_reset:
                            Gruppe.objects.all().delete()
                        wss = wb['Gruppen']
                        for row in wss.iter_rows(min_row=2, values_only=True):              # Erstellung Nutzer
                            #print(row)c
                            error = False
                            name, gruppenleiter_leiter, raum = row
                            if(Gruppe.objects.filter(name=name).exists()==False):
                                try:
                                    gruppenleiter_leiter = Personal.objects.get(user=(User.objects.get(username=gruppenleiter_leiter)))
                                except:
                                    error=True
                                    messages.error(request, fehler_tabelle+"gruppen_leiter "+gruppenleiter_leiter+" in Zeile " + row.count +" existiert nicht.")
                                try:
                                    raum = Raum.objects.get(raum_nr=raum)
                                except:
                                    error=True
                                    messages.error(request, fehler_tabelle+"raum "+raum+" in Zeile " + row.count +" existiert nicht.")
                                if not error:
                                    Gruppe.objects.create(name=name,gruppenleiter_leiter=gruppenleiter_leiter,raum=raum)
                            else:
                                messages.error(request, fehler_tabelle+"Gruppe "+name+" in Zeile " + row.count +" existiert bereits.")

                    if 'option_pupil' in optionlist:
                        fehler_tabelle='Fehler in Tabelle Schueler: '
                        if 'option_overwrite' in optionlist_reset:
                            Schueler.objects.all().delete()
                        wss = wb['Schueler']
                        for row in wss.iter_rows(min_row=2, values_only=True):              # Erstellung Nutzer
                            
                            error = False
                            vorname, nachname, gruppen_name, klasse, bus_kind, name_eb, kontakt_eb  = row
                            if(bus_kind.lower()=='true' or bus_kind.lower()=='ja'):
                                    is_offene_AG = True
                            else:
                                bus_kind = False
                            try:
                                gruppen_id = Gruppe.objects.get(name=gruppen_name)
                            except:
                                error=True
                                messages.error(request, fehler_tabelle+"Gruppe "+gruppen_id+" in Zeile " + row.count +" existiert nicht.")
                            new_nutzer = Nutzer.objects.create(vorname=vorname,nachname=nachname)
                            Schueler.objects.create(klasse=klasse, bus_kind=bus_kind, name_eb=name_eb, kontakt_eb=kontakt_eb, user_id=new_nutzer,gruppen_id=gruppen_id)
 
                    return redirect('home')
                    # return response                              # Site after Download
                # return response           # Weiterleitung wenn alles funktioniert hat
            except:
                messages.error(request, "Fehler bei importieren der Excel datei. Bitte überprüfe ob die Datei als 'Excel Arbeitsmappe (*.xlsx)' gespeichter ist, alle Tabellen die importier oder zurückgesetzt werden sollen existieren und ob die Spalten, wie in der Vorlage, richtig bennant sind.")              
                return redirect('home')      # Weiterleitung bei falscher xlsx datei
                                  
        print('Falscher Dateityp')        
        return render(request, 'csv_import/csv_import.html')      # Weiterleitung bei flaschen datei Typen   
    return render(request, 'csv_import/csv_import.html')